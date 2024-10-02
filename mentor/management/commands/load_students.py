import pandas as pd
from openpyxl import load_workbook
from django.core.management.base import BaseCommand
from ...models import MentorshipData  # Your model

class Command(BaseCommand):
    help = 'Load student data from Excel into the database'

    def handle(self, *args, **kwargs):
        # Load the Excel file
        file_path = 'SE_comps_Mentor_list_Final.xlsx'

        # Use openpyxl to open the workbook and get all sheets (including hidden ones)
        wb = load_workbook(file_path, data_only=True)

        # Iterate over each sheet, but process only visible ones
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]

            # Check if the sheet is visible
            if sheet.sheet_state == "visible":
                print(f"\nProcessing visible sheet: {sheet_name}")
                
                # Load the sheet using pandas, skip header rows
                df = pd.read_excel(file_path, sheet_name=sheet_name, skiprows=11)

                # Rename the columns (exclude 'Sr. No.')
                df.columns = ['Sr. No.', 'Name', 'Roll Number', 'Div', 'Faculty Mentor', 'BE Student Mentor']

                # Forward fill to handle any merged cells
                df.ffill(axis=0, inplace=True)

                # Drop any completely empty rows or columns
                df_cleaned = df.dropna(how='all', axis=1).dropna(how='all')

                # Drop the 'Sr. No.' column if it's not needed
                df_cleaned = df_cleaned.drop(columns=['Sr. No.'])

                # Replace remaining NaN values with 'Not Assigned'
                df_cleaned.fillna("Not Assigned", inplace=True)

                # Loop through each row of the cleaned DataFrame and save to the database
                for index, row in df_cleaned.iterrows():
                    MentorshipData.objects.create(
                        name=row['Name'],
                        roll_number=row['Roll Number'],
                        division=row['Div'],
                        faculty_mentor=row['Faculty Mentor'],
                        be_student_mentor=row['BE Student Mentor']
                    )
                print(f"Data from {sheet_name} loaded into the database.")

            else:
                print(f"Skipping hidden sheet: {sheet_name}")
